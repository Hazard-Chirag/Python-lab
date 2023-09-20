from openpyxl import Workbook
wb=Workbook()
wb.create_sheet()
sheets=wb.active
sheets.title="Language"
lang=['cat ','B','C']
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code =['KA', 'TS', 'TN']
sheets.cell(row=1,column=1).value="State"
sheets.cell(row=1,column=2).value="Language"
sheets.cell(row=1,column=3).value="Code"
for i in range(2,5):
    sheets.cell(row = i, column = 1).value = state[i-2]
    sheets.cell(row = i, column = 2).value = lang[i-2]
    sheets.cell(row = i, column = 3).value = code[i-2]
code=input("Etner the code to search: ")
wb.save("dog.xlsx")
for i in range(2,5):
    data=sheets.cell(row=i,column=3).value
    if(data==code):
        print(sheets.cell(row=i,column=2).value)
wb.close()